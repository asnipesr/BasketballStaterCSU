##TO DO FEATURE PLAN

# Calulcate Advanced Metrics -- create a similar one for regular statistics
# make it so that it can append on to existing file
# give rest of players 0 if not clicked
# make into excel file that looks similar to one existing
# Auto Save
# Live Feed
# Aggregate Stats Generator
# import csv and work from there in the case of a crash
# https://stackoverflow.com/questions/10020885/creating-a-popup-message-box-with-an-entry-field

import pygame
import pygame.locals 
from button import button, text
from datetime import date
import csv, os
import tkinter as tk
from tkinter import simpledialog, messagebox
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Fill, Border, Side
from pathlib import Path 
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell

# EXCEL 

ROOT = tk.Tk()
ROOT.withdraw()
CONTINUE_WB = False

# Gets wanted filename
today = date.today()
USER_INP = simpledialog.askstring(title="User Input", prompt="Name of File (don't include extension)")
if not USER_INP or USER_INP.strip() == "":
    base_name = f"./{today.month}-{today.day}statsheet"
else:
    base_name = USER_INP.strip()
    
file_name = base_name + ".xlsx"

# Checks if filename exists
# OPTIONS:
## Continue Working
## Add new worksheet
## Overwrite existing
if Path(file_name).is_file():
    ADDSHEET_INP = messagebox.askyesno(title="Pick Either", message="This file already exists. Do you want to add a new sheet to the workbook")

    if not ADDSHEET_INP:
        APPEND_INP = messagebox.askyesno(title="Continue Working", message="Do you want to continue working on an existing workbook?")
        
        if not APPEND_INP:
            OVERWRITE_INP = messagebox.askyesno(title="Pick one", message="Are you sure? Selecting YES will overwrite existing files! Selecting NO will create a copy")
            if not OVERWRITE_INP:
                count = 0
                if not OVERWRITE_INP:   
                    while Path(file_name).is_file():
                        count+=1
                        file_name = f"./{base_name}_{count}.xlsx"
    else:
        CONTINUE_WB = True
        wb = load_workbook(filename=file_name)
        SHEETNAME_INP = simpledialog.askstring(title="New Sheet Input", prompt="Enter name of name of the new sheet (can't use '/' character)")
        
        if not SHEETNAME_INP or SHEETNAME_INP.strip() == "":
            sheet_name = f"./{today.month}-{today.day}statsheet"
        else:
            sheet_name = SHEETNAME_INP.strip()
        
        # same thing with checking to append or overwrite sheet
        
        count = 1
        while sheet_name in wb.sheetnames:
            sheet_name = f"{sheet_name} {count}"
            count+=1  

# check it out
print("Saving name", file_name)
print("\n")
    
#initializing pygame
pygame.init()

#setting screen width and height
SCREEN_WIDTH = 800
SCREEN_HEIGHT = 800
ROSTER_SIZE = 15

#creating game window
screen = pygame.display.set_mode((SCREEN_WIDTH, SCREEN_HEIGHT))
pygame.display.set_caption("Stat Tracker")

#creating text for title
screen.fill((0,0,0))
pygame.font.init()
textfont = pygame.font.SysFont("monospace", 50)

textTBR = textfont.render("STAT TRACKER", 1, (255,255,255))
screen.blit(textTBR, (220, 10))

# create a surface object, image is drawn on it.
imp = pygame.image.load("csuLogo.webp").convert()
imp = pygame.transform.scale(imp,(250,250))
 
# Using blit to copy content from one surface to other
screen.blit(imp, (275, 50))
 
# paint screen one time
pygame.display.flip()

run = True
selected = False
global curr

def generate_player_rows(stats: dict):
    multipliers = [3,-1,2,-1,1,-2,1,2,2,-3,1,2,1,0,2,2,1,3,1,-1,-1,1,1,-1,0]
    player_rows = []
    real_data = []
    for person in stats.keys():
        values = stats[person] 
        total = sum(int(a) * b for a, b in zip(values, multipliers))
        row_data = [person] + values[:25] + [total] 
        player_rows.append(row_data)
        real_data.append(([person] + values))
        
    # Generates row for player if they are not already on the board
    for player in players:
        name = player["name"]
        if name not in stats:
            zero_stats = [0] * 30
            total = 0 
            row_data = [name] + zero_stats[:25] + [total]
            player_rows.append(row_data)
            real_data.append(([name] + zero_stats))
    player_rows.sort(key=lambda row:row[0])
    real_data.sort(key=lambda row:row[0])
    return player_rows, real_data

import tkinter as tk
from tkinter import ttk
import pygame

# ---- position & modality helpers ----
def _center_tk_on_screen(top, width, height):
    top.update_idletasks()
    sw = top.winfo_screenwidth()
    sh = top.winfo_screenheight()
    x = (sw - width)//2
    y = (sh - height)//2
    top.geometry(f"{width}x{height}+{x}+{y}")


def _make_modal(top: tk.Toplevel):
    top.attributes("-topmost", True)  # stay above pygame window
    top.grab_set()                    # modal: block rest until closed
    top.focus_force()
    top.transient()                   # hint "child of parent app"

# ---- reusable choice dialog (2 buttons) ----
def ask_choice_tk(title: str,
                  prompt: str,
                  left_text: str,
                  right_text: str,
                  left_value,
                  right_value,
                  *,
                  width=260, height=140,
                  bg="#0B0B0B",           # almost-black (matches your app)
                  fg="#FFFFFF",           # white text
                  accent="#1B6A42"        # CSU green
                  ):
    """
    Returns left_value or right_value, or None if closed/esc.
    """
    top = tk.Toplevel()
    top.withdraw()  # avoid flicker while we style/place
    top.title(title)
    top.configure(bg=bg)

    # center over pygame window
    _center_tk_on_screen(top, width, height)

    # simple theming
    style = ttk.Style(top)
    style.theme_use("clam")
    style.configure("Dark.TLabel", background=bg, foreground=fg, font=("Oswald", 11, "bold"))
    style.configure("Dark.TButton", font=("Oswald", 10, "bold"), padding=6)
    style.map("Dark.TButton",
              background=[("active", accent)],
              foreground=[("active", "#FFFFFF")])

    # layout
    frm = ttk.Frame(top, padding=10)
    frm.pack(expand=True, fill="both")
    frm.configure(style="Dark.TLabel")  # inherit bg

    lbl = ttk.Label(frm, text=prompt, style="Dark.TLabel")
    lbl.pack(pady=(4, 10))

    btn_row = ttk.Frame(frm)
    btn_row.pack()

    result = {"val": None}

    def _choose(v):
        result["val"] = v
        top.destroy()

    b1 = ttk.Button(btn_row, text=left_text, style="Dark.TButton", command=lambda: _choose(left_value))
    b2 = ttk.Button(btn_row, text=right_text, style="Dark.TButton", command=lambda: _choose(right_value))
    b1.pack(side="left", padx=8)
    b2.pack(side="left", padx=8)

    # key bindings
    top.bind("<Escape>", lambda e: top.destroy())
    top.bind("2",       lambda e: _choose(left_value))
    top.bind("3",       lambda e: _choose(right_value))
    top.bind("<Return>",lambda e: _choose(right_value))  # optional default

    # show & make modal
    top.deiconify()
    _make_modal(top)
    top.wait_window()
    
    pygame.display.get_wm_info()
    pygame.event.pump()
    return result["val"]

def ask_gold_input():
    return ask_choice_tk(
        title="Specify Gold Type",
        prompt="Select Gold Shot Type:",
        left_text="Gold 2",  right_text="Gold 3",
        left_value=2,        right_value=3
    )

def ask_ft_input():
    return ask_choice_tk(
        title="Specify FT Type",
        prompt="Select FT Type:",
        left_text="FT Miss",  right_text="FT Make",
        left_value=0,        right_value=1
    )
   
# Sends stats to file and formats worksheet
def send_to_file(stats, wb=None, sheet_name=""):
    header = ["PLAYER","GOLD\n +3", "GOLD MISS\n -1", "SILVER\n +2", "SILVER MISS\n -1","BRONZE\n +1", "BRONZE MISS\n -2", "FTS\n +1", "AST\n +2", "VIKING AST\n +2", "TO\n -3", "PT\n +1", "OREB\n +2", "DREB\n +1", "REB", "STL\n +2", "BLK\n +2", "DEFL\n +1", "CHG\nW-UP\n +3", "DRAW FL\n +1", "FOUL\n -1", "BLOW BY\n -1", "WIN\n +1", "GOOD CUT\n+1", "BAD CUT\n-1", "POSS", "TOTAL"]
    
    # Create workbook if none exists (doesn't want to append)
    if wb is None:
        wb = Workbook()
        ws = wb.active
        if ws.title == "Sheet":
            wb.remove(ws)
    
    # Clear cells in worksheet or creates new one
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if not isinstance(cell, MergedCell):
                    cell.value = None
    else:
        ws = wb.create_sheet(title=sheet_name)
    
    merged_column = get_column_letter(len(header))
    # Generate Top Header
    ws.merge_cells(f"A1:{merged_column}1")
    ws['A1'] = "Cleveland State Basketball"
    ws['A1'].font = Font(name=" Oswald", size=22, bold=True, color=("FFFFFF"))
    ws['A1'].fill = PatternFill(start_color="1B6A42", end_color="1B6A42",fill_type="solid")
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    
    # Generate secondary header
    ws.merge_cells(f"A2:{merged_column}2")
    ws['A2'] = f"Viking Way Stats - {today.month}/{today.day}"
    ws['A2'].font = Font(name="Oswald", size=16, italic=True, color="000000")
    ws['A2'].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9",fill_type="solid")
    ws['A2'].alignment = Alignment(horizontal="center", vertical="center")
    
    # Create header with wanted statistics
    header_font = Font(name="Oswald", bold=True, italic=True, color="FFFFFF")
    header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    ## Draws statistic headers
    header_row = 3
    for col_index, value in enumerate(header, start=1):
        cell = ws.cell(row=header_row, column=col_index, value=value)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        
    # Generates rows of player statistic data
    player_rows, total_rows = generate_player_rows(stats)
    
    start_row = 4
    for r_index, row in enumerate(player_rows, start=start_row):
        for c_index, value in enumerate(row, start=1):
            ws.cell(row=r_index, column=c_index, value=value)   
    
    # Styling for Rows and text
    for r in range(4, 4+ROSTER_SIZE):
        ws[f"A{r}"].font = Font(name="Oswald", size=12)
        ws[f"A{r}"].alignment = Alignment(vertical="center")
        ws[f"{merged_column}{r}"].font = Font(size=12, name="Oswald", bold=True)
        ws[f"{merged_column}{r}"].alignment = Alignment(horizontal="center", vertical="center") 
        ws[f"{merged_column}{r}"].border = Border(left=Side(style="thick", color="000000"), bottom=Side(style="thin", color="000000"), right=Side(style="thin", color="000000"))
        ws.row_dimensions[r].height = 22.00     
    
    for r in range(4, 4+ROSTER_SIZE):
        for c in range(1, len(header)):
            cell = ws.cell(row=r, column=c)
            if c == 1:
                cell.border = Border(right=Side(style="thin", color="000000"), bottom=Side(style="thin", color="000000")) 
                continue
           
            cell.font = Font(name="Oswald", size=12)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(right=Side(style="thin", color="000000"), bottom=Side(style="thin", color="000000")) 

    for r in range(5, 4+ROSTER_SIZE, 2):
        for c in range(1, len(header)+1):
            cell = ws.cell(row=r, column=c)
            cell.fill = PatternFill(start_color="efefef", end_color="efefef", fill_type="solid")
    
    """ for r in range(4, 4+ROSTER_SIZE):
        for c in range(1, len(header)):       
            cell = ws.cell(row=r, column=c) 
            cell.border = Border(right=Side(style="thin", color="000000"), bottom=Side(style="thin", color="000000"))  """
            
    ws.merge_cells(f"A{ROSTER_SIZE+4}:{merged_column}{ROSTER_SIZE+4}")
    ws[f"A{ROSTER_SIZE+4}"].fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    ws.row_dimensions[4 + ROSTER_SIZE].height = 26.00
    ws.row_dimensions[3].height = 57.00
            
    # Spaces cells width and height correctly 
    OFFSET = 0.83
    for col in ws.columns:
        column = get_column_letter(col[0].column)  # Convert 1 -> 'A', etc.
        if column == 'A':
            ws.column_dimensions[column].width = 16.00 + OFFSET
        elif column == 'O':
            ws.column_dimensions[column].width = 5.00 + OFFSET
        else:
            ws.column_dimensions[column].width = 7.50 + OFFSET
            
    header_hidden = ["Gold 2 Make", "Gold 2 Miss", "Gold 3 Make", "Gold 3 Miss", "FT Misses"]
    hidden_players = []
    for person in total_rows:
        #name = person[0]
        data = person[26:]
        hidden_players.append(data)
    
    header_row = 3
    for col_index, value in enumerate(header_hidden, start=31):
        cell = ws.cell(row=header_row, column=col_index, value=value)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        
    for r_index, row in enumerate(hidden_players, start=start_row):
        for c_index, value in enumerate(row, start=31):
            ws.cell(row=r_index, column=c_index, value=value) 
            
    for r in range(4, 4+ROSTER_SIZE):
        for c in range(31, 31+len(header_hidden)):
            cell = ws.cell(row=r, column=c)
            cell.font = Font(name="Oswald", size=12)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(left=Side(style="thin", color="000000"), right=Side(style="thin", color="000000"), 
                                 bottom=Side(style="thin", color="000000")) 

    for r in range(5, 4+ROSTER_SIZE, 2):
        for c in range(31, 31+len(header_hidden)):
            cell = ws.cell(row=r, column=c)
            cell.fill = PatternFill(start_color="efefef", end_color="efefef", fill_type="solid")
    
    
    wb.save(file_name)   
    
    
# "WIN" ,"FGM", "FGA", "3PM", "3PA", "AST", "ORB", "DRB","STL","BLK","TOV"]
# "GOLD", "SILVER", "BRONZE", "FTS", "AST", "TO", "PT", "OREB", "DREB", "REB", "STL", "BLK", "DEFL", "CHG/W-UP", "DRAW FL", "FOUL", "BLOW BY" "TEAM WIN", 
def GOLD():
    global curr
    curr = "GOLD"
def GOLD_MISS():
    global curr
    curr = "GOLD MISS"
def SILVER():
    global curr
    curr = "SILVER"
def SILVER_MiSS():
    global curr
    curr = "SILVER MISS"
def BRONZE():
    global curr
    curr = "BRONZE"
def BRONZE_MISS():
    global curr
    curr = "BRONZE MISS"
def FTs():
    global curr
    curr = "FTs"
def AST():
    global curr
    curr = "AST"
def Viking_AST():
    global curr
    curr = "Viking_AST"
def TOV():
    global curr
    curr = "TOV"
def PT():
    global curr
    curr = "PT"
def OREB():
    global curr
    curr = "OREB"
def DREB():
    global curr
    curr = "DREB"
def STL():
    global curr
    curr = "STL"
def BLK():
    global curr
    curr = "BLK"
def DEFL():
    global curr
    curr = "DEFL"
def CHG_WUP():
    global curr
    curr = "CHG/W-UP"
def DRAW_FL():
    global curr
    curr = "DRAW FOUL"
def FOUL():
    global curr
    curr = "FOUL"
def BLOW_BY():
    global curr
    curr = "BLOW BY"
def WIN():
    global curr
    curr = "TEAM WIN"
def GOOD_CUT():
    global curr
    curr = "GOOD CUT"
def BAD_CUT():
    global curr
    curr = "BAD CUT"
def POSS():
    global curr
    curr = "POSS"


def find(lst, num):
    for dictionary in lst:
        if dictionary["number"] == num:
            return dictionary
    return {}

def find_option(lst, name):
    for dictionary in lst:
        if dictionary["name"] == name:
            return dictionary
    return {}

stats = {}
global stat_records
stat_records = []
def Number(num):
    global stat_records
    new_stats = {}
    for i in stats.keys():
        new_stats[i] = stats[i].copy()
    stat_records.append(new_stats)
    player_dict = find(players, num)
    name = player_dict["name"]
    stats_dict = find_option(options, curr)
    
    gold_id = 0
    if curr == "GOLD" or curr == "GOLD MISS":
        gold_id = ask_gold_input()
        if name not in stats:
            stats[name] = [0]*30
            
        if gold_id == 2 and curr == "GOLD":
            stats[name][25] += 1
            print("GOLD 2 MAKE")
            
        elif gold_id == 2 and curr == "GOLD MISS":
            stats[name][26] += 1
            print("GOLD 2 MISS")
            
        elif gold_id == 3 and curr == "GOLD":
            stats[name][27] += 1
            print("GOLD 3 MAKE")
            
        elif gold_id == 3 and curr == "GOLD MISS":
            stats[name][28] += 1
            print("GOLD 3 MISS")
          
    if curr == "POSS":
        poss_val = None
        try:
            while poss_val is None:
                POSS_INP = simpledialog.askinteger(title=f"{stats_dict["name"]}", prompt=f"Enter number of {stats_dict["name"]} for {name}", minvalue=0, maxvalue=200)
                if POSS_INP is None:
                    return
                poss_val = POSS_INP
        except Exception:
            messagebox.showerror("Invalid Entry. Please enter a number between 0 and 200")
        
        if name not in stats:
            stats[name] = [0]*30
        
        stats[name][stats_dict["index"]] = poss_val
        print(name + " -- " + "Possession value: " + str(poss_val))
    
    elif curr == "FTs":
        ft_id = 0
        ft_id = ask_ft_input()
        if name not in stats:
            stats[name] = [0] * 30
            
        if ft_id == 0:
            stats[name][29] += 1
            print(name + " -- " + "FT MISS")
        else:
            stats[name][stats_dict["index"]] += 1
            print(name + " -- " + "FT MAKE")
    else:
        if name in stats:
            stats[name][stats_dict["index"]] += 1
            if stats_dict["index"] == 11 or stats_dict["index"] == 12:
                stats[name][13] += 1
        else:
            stats[name] = [0]*30
            stats[name][stats_dict["index"]] = 1
            if stats_dict["index"] == 11 or stats_dict["index"] == 12:
                stats[name][13] += 1
                 
        print(name + " -- " + stats_dict["name"])
    
    pass

players = [
    {"number": "0",
     "function": Number,
     "name": "Foster Wonders",
     "img": "./players/0foster.webp"},
    {"number": "1",
     "function": Number,
     "name": "Ice Emery Jr.",
     "img": "./players/1ice.webp"},
    {"number": "2",
     "function": Number,
     "name": "Jaidon Lipscomb",
     "img": "./players/2jaidon.webp"},
    {"number": "3",
     "function": Number,
     "name": "Tre Beard",
     "img": "./players/3tre.webp"},
    {"number": "4",
     "function": Number,
     "name": "Preist Ryan",
     "img": "./players/4preist.webp"},
    {"number": "5",
     "function": Number,
     "name": "David Giddens",
     "img": "./players/5david.webp"},
    {"number": "7",
     "function": Number,
     "name": "Dayan Nessah",
     "img": "./players/7dayan.png"},
    {"number": "11",
     "function": Number,
     "name": "Waqo Tessema",
     "img": "./players/11waqo.webp"},
    {"number": "12",
     "function": Number,
     "name": "Holden Pierre-Louis",
     "img": "./players/12holden.webp"},
    {"number": "13",
     "function": Number,
     "name": "Lucas Burton",
     "img": "./players/13lucas.webp"},
    {"number": "15",
     "function": Number,
     "name": "Ivan Spirov",
     "img": "./players/15ivan.png"},
    {"number": "22",
     "function": Number,
     "name": "Josiah Harris",
     "img": "./players/22jojo.webp"},
    {"number": "23",
     "function": Number,
     "name": "Manny Hill",
     "img": "./players/23manny.webp"},
    {"number": "24",
     "function": Number,
     "name": "Kamari Jones",
     "img": "./players/24kamari.webp"},
    {"number": "50",
     "function": Number,
     "name": "Kevo St. Hilaire",
     "img": "./players/50kevo.webp"},
]
# "GOLD", "SILVER", "BRONZE", "FTS", "AST", "TO", "PT", "OREB", "DREB", "REB", "STL", "BLK", "DEFL", "CHG/W-UP", "DRAW FL", "FOUL", "BLOW BY" "TEAM WIN", 

options = [
    { "name" : "GOLD",
      "function": GOLD,
      "index": 0
    },
    { "name" : "GOLD MISS",
      "function": GOLD_MISS,
      "index": 1
    },
    { "name" : "SILVER",
      "function": SILVER,
      "index": 2
    },
    { "name" : "SILVER MISS",
      "function": SILVER_MiSS,
      "index": 3
    },
    { "name" : "BRONZE",
      "function": BRONZE,
      "index": 4
    },
    { "name" : "BRONZE MISS",
      "function": BRONZE_MISS,
      "index": 5
    },
    { "name" : "FTs",
      "function": FTs,
      "index": 6
    },
    { "name" : "AST",
      "function": AST,
      "index": 7
    },
    { "name" : "Viking_AST",
      "function": Viking_AST,
      "index": 8
    },
    { "name" : "TOV",
      "function": TOV,
      "index": 9
    },
    { "name" : "PT",
      "function": PT,
      "index": 10
    },
    { "name" : "OREB",
      "function": OREB,
      "index": 11
    },
    { "name" : "DREB",
      "function": DREB,
      "index": 12
    },
    #total rebounds is index 13
    { "name" : "STL",
      "function": STL,
      "index": 14
    },
    { "name" : "BLK",
      "function": BLK,
      "index": 15
    },
    { "name" : "DEFL",
      "function": DEFL,
      "index": 16
    },
    { "name" : "CHG/W-UP",
      "function": CHG_WUP,
      "index": 17
    },
    { "name" : "DRAW FOUL",
      "function": DRAW_FL,
      "index": 18
    },
    { "name" : "FOUL",
      "function": FOUL,
      "index": 19
    },
    { "name" : "BLOW BY",
      "function": BLOW_BY,
      "index": 20
    },
    { "name" : "TEAM WIN",
      "function": WIN,
      "index": 21
    },
    { "name" : "GOOD CUT",
      "function": GOOD_CUT,
      "index": 22
    },
    { "name" : "BAD CUT",
      "function": BAD_CUT,
      "index": 23
    },
    { "name" : "POSS",
     "function": POSS,
     "index": 24
    }
]

button_list = []
x = 80
x_increment = 125
y = 425
row = 0
y_increment = 75
count = 0

# Creates initial buttons - Statistics
for option in options:
    pos_x = x + (count * x_increment)
    pos_y = y + (row * y_increment)
    
    curr_button = button(position = (pos_x, pos_y), size=(100, 50), clr=(220, 220, 220), cngclr=(255, 0, 0), func=option["function"], text=option["name"])
    count+=1
    
    if pos_x > 700:
        count = 0
        row += 1
        
    button_list.append(curr_button)

player_list = []
player_images = []
x2 = 80
x_increment2 = 125
y2 = 425
row = 0
y_increment2 = 160
count = 0

# Create player buttons
for option in players:
    if "img" in players:
        img = pygame.image.load(option["img"]).convert_alpha()
        img = pygame.transform.scale(img, (100,125))
        option["img"] = img

for option in players:
    pos_x = x2 + (count * x_increment2)
    pos_y = y2 + (row * y_increment2)
    
    curr_button = button(position = (pos_x, pos_y), size=(100, 50), clr=(220, 220, 220), cngclr=(255, 0, 0), func=option["function"], text=option["number"])
    count+=1
    
    if pos_x > 700:
        count = 0
        row += 1
    player_list.append(curr_button)
    
    # Places image to corresponding player 
    if "img" in option:
        img = pygame.image.load(option["img"]).convert_alpha()
        img = pygame.transform.scale(img, (70,87))
        img_rect = img.get_rect(center=(pos_x, pos_y - 73))
        # player_images.append((img, img_rect))
        player_images.append({"img": img, "rect": img_rect, "player_number": option["number"], "func": option["function"]})

def save():
    if stats == {}:
        print("No Stats To Save")
    else:
        if CONTINUE_WB:
            send_to_file(stats,wb=wb,sheet_name=sheet_name)
        else:
            send_to_file(stats,sheet_name="Stat Sheet")
        print("\nSaved to ", file_name, "\nPath: ", os.getcwd())

def new_game():
    stats = {}
    players = []

while run:
    
    for event in pygame.event.get():
        if event.type == pygame.QUIT:
            save()
            run = False
        elif event.type == pygame.KEYDOWN:
            if pygame.key.get_mods() & pygame.KMOD_META:
                if event.key == pygame.K_s:
                    save()   
                if event.key == pygame.K_z:
                    print("Undo")
                    if len(stat_records) > 0:
                        stats = stat_records[len(stat_records)-1].copy()
                        stat_records = stat_records[0:-1]
                        #print(stats)
                if event.key == pygame.K_1:
                    quit
        elif event.type == pygame.MOUSEBUTTONDOWN:
                if event.button == 1:
                    pos = pygame.mouse.get_pos()
                    pygame.draw.rect(screen, (0,0,0), pygame.Rect(0, 300, 800, 500))
                    if not selected and not press:
                        for b in button_list:
                            if b.rect.collidepoint(pos):
                                selected = True
                                press = True
                                pygame.draw.rect(screen, (0,0,0), pygame.Rect(0, 400, 800, 200))
                                pygame.display.flip()
                                b.call_back()
                    if selected and not press:
                        for b in player_list:
                            if b.rect.collidepoint(pos):
                                selected = False
                                press = True
                                pygame.draw.rect(screen, (0,0,0), pygame.Rect(0, 400, 800, 200))
                                pygame.display.flip()
                                b.call_back(b.txt)
                        for p in player_images:
                            if p["rect"].collidepoint(pos):
                                selected = False
                                press = True
                                pygame.draw.rect(screen, (0,0,0), pygame.Rect(0, 400, 800, 200))
                                pygame.display.flip()
                                p["func"](p["player_number"])
        else:
            press = False
                
    if not selected:
        for b in button_list:
            b.draw(screen)
    if selected:
        for b in player_list:
            b.draw(screen)
        for p in player_images:
            screen.blit(p["img"], p["rect"])

    pygame.display.update()

pygame.quit()